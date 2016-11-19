# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.utils import rows_from_range
from prettyprint import pp
from functools import partial
def getLevelTypes(ws, tl, rb):
    return map(lambda x: x.value, ws[tl: rb][0])

def getValueWithMergeLookup(sheet, cell):
    idx = cell.coordinate
    for range_ in sheet.merged_cell_ranges:
        merged_cells = list(rows_from_range(range_))
        for row in merged_cells:
            if idx in row:
                # If this is a merged cell,
                # return  the first cell of the merge range
                return sheet[merged_cells[0][0]].value

    return sheet.cell(idx).value

def getMergedCellPresent(sheet, cell):
    idx = cell.coordinate
    for range_ in sheet.merged_cell_ranges:
        merged_cells = list(rows_from_range(range_))
        for row in merged_cells:
            if idx in row:
                return sheet[merged_cells[0][0]]
    return cell

def genLevel(ws, remains, callback):
    cellAndValues = map(lambda x: (x, getMergedCellPresent(ws, x).value), remains)
    callback(cellAndValues)

def pushLevels(levelDict, cellAndValues):
    currentDict = levelDict
    for cell, value in cellAndValues:
        if value not in currentDict:
            currentDict[value] = {}
        currentDict = currentDict[value]

def travesalLevels(ws, tl, rb, onCell):
    levels = ws[tl : rb]
    for row in levels:
        genLevel(ws, row, onCell)

if __name__ == '__main__':
    print 'here'
    import sys
    wb = load_workbook(filename = "../test.xlsx", read_only = False)
    ws = wb.worksheets[0]
    levelTypes = getLevelTypes(ws, 'B6', 'D7')
    levelDict = {}
    travesalLevels(ws, 'B8', 'D419', partial(pushLevels, levelDict))
    pp(levelDict)
